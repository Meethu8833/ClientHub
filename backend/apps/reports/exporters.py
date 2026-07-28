"""
File exporters (docs/reports-module.md): ReportTable → downloadable bytes.

Both exporters are dumb on purpose. They know NOTHING about invoices or
tickets — only the ReportTable contract (columns with a `kind`, rows,
totals). Add a fourth report and these functions already export it.

Both build the file in memory (BytesIO). Fine here: reports are aggregated
rows (a year of months = 12 rows, clients = hundreds), never the raw tables.
If a report ever returns 100k+ rows, switch to a StreamingHttpResponse and
openpyxl's write_only mode — and question the report design first.

Security note (the classic export bug): a cell that STARTS with = + - or @
is executed as a formula by Excel. A client named "=HYPERLINK(...)" typed
into the CRM would detonate in the accountant's spreadsheet. _safe_text()
neutralises it with a leading apostrophe — Excel's own "treat as text"
escape. PDF has the equivalent trap: reportlab Paragraphs parse XML-ish
markup, so all text is XML-escaped before it goes in.
"""

from decimal import Decimal
from io import BytesIO
from xml.sax.saxutils import escape

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from .services import Column, ReportTable

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
PDF_MIME = "application/pdf"

# Excel number-format strings per column kind.
NUMBER_FORMATS = {"money": "#,##0.00", "hours": "0.00", "int": "0"}


def _safe_text(value: str) -> str:
    """Defuse spreadsheet formula injection: '=SUM(A1)' → \"'=SUM(A1)\"."""
    if value and value[0] in "=+-@":
        return "'" + value
    return value


def _filters_line(table: ReportTable) -> str:
    """'date_from: 2025-07-28 · group_by: month' — the report's fine print."""
    parts = [f"{k}: {v}" for k, v in table.filters.items() if v is not None]
    return " · ".join(parts)


def _attachment(payload: bytes, mime: str, filename: str) -> HttpResponse:
    """
    Content-Disposition: attachment is what makes the browser SAVE the bytes
    instead of trying to render them; filename= is what the save dialog shows.
    """
    response = HttpResponse(payload, content_type=mime)
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


# ---------------------------------------------------------------------------
# Excel (openpyxl)
# ---------------------------------------------------------------------------


def xlsx_response(table: ReportTable, filename: str) -> HttpResponse:
    wb = Workbook()
    ws = wb.active
    ws.title = table.title[:31]  # Excel hard-caps sheet names at 31 chars

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F2937")  # slate-800
    totals_font = Font(bold=True)
    totals_border = Border(top=Side(style="thin"))

    # Row 1: title. Row 2: the filters that produced these numbers — an
    # exported file gets emailed around; it must carry its own context.
    ws.cell(row=1, column=1, value=table.title).font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=_filters_line(table)).font = Font(italic=True, color="6B7280")

    header_row = 4
    for col_idx, col in enumerate(table.columns, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=col.label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="right" if col.kind != "text" else "left")
    # Freeze everything above row 5 so headers stay visible while scrolling.
    ws.freeze_panes = f"A{header_row + 1}"

    data_rows = list(table.rows)
    if table.totals:
        data_rows.append(table.totals)
    for r_off, row in enumerate(data_rows):
        is_totals = table.totals is not None and r_off == len(data_rows) - 1
        for col_idx, col in enumerate(table.columns, start=1):
            value = row.get(col.key)
            if isinstance(value, str):
                value = _safe_text(value)
            elif isinstance(value, Decimal):
                value = float(value)  # native number → cell stays summable
            cell = ws.cell(row=header_row + 1 + r_off, column=col_idx, value=value)
            if col.kind in NUMBER_FORMATS:
                cell.number_format = NUMBER_FORMATS[col.kind]
            if is_totals:
                cell.font = totals_font
                cell.border = totals_border

    # Width ≈ widest content in the column, clamped to sane bounds.
    for col_idx, col in enumerate(table.columns, start=1):
        widest = max([len(col.label)] + [len(str(r.get(col.key) or "")) for r in data_rows] or [0])
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(widest + 3, 10), 40)

    buffer = BytesIO()
    wb.save(buffer)
    return _attachment(buffer.getvalue(), XLSX_MIME, filename)


# ---------------------------------------------------------------------------
# PDF (reportlab / platypus)
# ---------------------------------------------------------------------------


def _pdf_cell(value, col: Column, style) -> object:
    """Format one value for print. Numbers become strings HERE (PDF is paper
    — nothing is summable anyway); text is escaped and wrapped."""
    if value is None:
        return "—"
    if isinstance(value, Decimal):
        return f"{value:,.2f}"
    if col.kind == "text":
        return Paragraph(escape(str(value)), style)
    return str(value)


def pdf_response(table: ReportTable, filename: str) -> HttpResponse:
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),  # reports are wide; portrait truncates
        title=table.title,
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
    )
    styles = getSampleStyleSheet()
    body_style = styles["BodyText"]

    grid = [[col.label for col in table.columns]]
    for row in table.rows:
        grid.append([_pdf_cell(row.get(c.key), c, body_style) for c in table.columns])
    if table.totals:
        grid.append([_pdf_cell(table.totals.get(c.key), c, body_style) for c in table.columns])

    style = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F2937")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#9CA3AF")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F3F4F6")]),
    ]
    # Right-align every numeric column, whole column at once.
    for idx, col in enumerate(table.columns):
        if col.kind != "text":
            style.append(("ALIGN", (idx, 0), (idx, -1), "RIGHT"))
    if table.totals:
        style += [
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
            ("LINEABOVE", (0, -1), (-1, -1), 0.75, colors.black),
            ("BACKGROUND", (0, -1), (-1, -1), colors.white),
        ]

    story = [
        Paragraph(escape(table.title), styles["Title"]),
        Paragraph(escape(_filters_line(table)), styles["Italic"]),
        Spacer(1, 0.5 * cm),
        # repeatRows=1 re-prints the header on every page of a long table.
        Table(grid, repeatRows=1, style=TableStyle(style)),
    ]
    doc.build(story)
    return _attachment(buffer.getvalue(), PDF_MIME, filename)
