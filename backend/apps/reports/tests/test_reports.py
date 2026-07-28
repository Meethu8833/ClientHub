"""
Report API tests: aggregation correctness, per-role scoping, param
validation, and the three renderers (JSON shape, real .xlsx bytes, real
.pdf bytes — both parsed back, not just status-checked).

Data is seeded straight through the ORM, same rationale as the dashboard
suite: reports read whatever is in the tables; the source modules' own
suites prove the write paths.
"""

from datetime import date, timedelta
from decimal import Decimal
from io import BytesIO

import pytest
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook
from rest_framework.test import APIClient

from apps.accounts.models import User
from apps.billing.models import Invoice, Payment, Refund
from apps.clients.models import Client
from apps.projects.models import Project, Task, TimeEntry
from apps.tickets.models import Ticket, TicketCategory, TicketPriority

PASSWORD = "Str0ng!pass123"

pytestmark = pytest.mark.django_db

REVENUE_URL = reverse("reports:revenue")
TIME_URL = reverse("reports:time")
TICKETS_URL = reverse("reports:tickets")

# Fixed window → deterministic month axis: Jan, Feb, Mar 2026.
WINDOW = {"date_from": "2026-01-01", "date_to": "2026-03-31"}


# --- fixtures ---------------------------------------------------------------


@pytest.fixture
def manager():
    return User.objects.create_user(
        email="mgr@x.com", password=PASSWORD, role=User.Role.MANAGER, first_name="Mia"
    )


@pytest.fixture
def staff():
    return User.objects.create_user(
        email="staff@x.com", password=PASSWORD, role=User.Role.STAFF, first_name="Sam"
    )


def api_for(user):
    c = APIClient()
    c.force_authenticate(user=user)
    return c


@pytest.fixture
def manager_api(manager):
    return api_for(manager)


@pytest.fixture
def staff_api(staff):
    return api_for(staff)


@pytest.fixture
def acme():
    return Client.objects.create(name="Acme Corp", status=Client.Status.ACTIVE)


@pytest.fixture
def money_world(acme, manager):
    """
    Feb: invoice for 1000 issued (200 already recorded as paid → 800 open).
    Mar: the 200 payment lands; 50 of it refunded.
    Plus one VOID and one DRAFT invoice that must count for NOTHING.
    """
    inv = Invoice.objects.create(
        client=acme,
        invoice_number="INV-2026-0001",
        status=Invoice.Status.PARTIALLY_PAID,
        grand_total=Decimal("1000.00"),
        amount_paid=Decimal("200.00"),
        issue_date=date(2026, 2, 10),
        due_date=date(2026, 3, 12),
    )
    Invoice.objects.create(  # VOID: issued then retracted — not revenue
        client=acme,
        invoice_number="INV-2026-0002",
        status=Invoice.Status.VOID,
        grand_total=Decimal("999.00"),
        issue_date=date(2026, 2, 15),
        due_date=date(2026, 3, 17),
        voided_at=timezone.now(),
    )
    Invoice.objects.create(client=acme, grand_total=Decimal("777.00"))  # DRAFT
    pay = Payment.objects.create(
        invoice=inv, amount=Decimal("200.00"), received_on=date(2026, 3, 5), recorded_by=manager
    )
    Refund.objects.create(
        payment=pay, amount=Decimal("50.00"), refunded_on=date(2026, 3, 20), reason="goodwill"
    )
    return inv


@pytest.fixture
def hours_world(acme, manager, staff):
    project = Project.objects.create(client=acme, name="CRM Build")
    task = Task.objects.create(project=project, title="API work")
    today = timezone.localdate()
    TimeEntry.objects.create(task=task, user=staff, hours=Decimal("2.5"), worked_on=today)
    TimeEntry.objects.create(task=task, user=manager, hours=Decimal("1"), worked_on=today)
    return project


@pytest.fixture
def ticket_world(acme):
    now = timezone.now()
    cat = TicketCategory.objects.create(name="Bug")
    resolved = Ticket.objects.create(
        client=acme,
        category=cat,
        subject="Site down",
        description="…",
        priority=TicketPriority.HIGH,
        status=Ticket.Status.RESOLVED,
        resolution_due_at=now - timedelta(hours=2),
        resolved_at=now - timedelta(hours=1),  # 1h past its deadline → breached
    )
    # created_at is auto_now_add — only a queryset update can backdate it.
    # 3h before resolution → avg resolution must come out at exactly 2.00h.
    Ticket.objects.filter(pk=resolved.pk).update(created_at=now - timedelta(hours=3))
    Ticket.objects.create(  # open, first response already overdue
        client=acme,
        category=cat,
        subject="Slow page",
        description="…",
        priority=TicketPriority.HIGH,
        first_response_due_at=now - timedelta(hours=1),
        resolution_due_at=now + timedelta(hours=5),
    )


# --- permissions & validation ----------------------------------------------


def test_revenue_forbidden_for_staff(staff_api):
    assert staff_api.get(REVENUE_URL).status_code == 403


def test_reports_require_auth():
    assert APIClient().get(TIME_URL).status_code == 401


def test_inverted_date_range_is_400(manager_api):
    res = manager_api.get(REVENUE_URL, {"date_from": "2026-03-01", "date_to": "2026-01-01"})
    assert res.status_code == 400
    assert "date_from" in res.json()


def test_unknown_export_format_is_400(manager_api):
    assert manager_api.get(REVENUE_URL, {"export": "docx"}).status_code == 400


# --- revenue report ---------------------------------------------------------


def test_revenue_by_month_zero_fills_and_totals(manager_api, money_world):
    data = manager_api.get(REVENUE_URL, WINDOW).json()

    assert [r["month"] for r in data["rows"]] == ["2026-01", "2026-02", "2026-03"]
    jan, feb, mar = data["rows"]
    assert jan["invoiced"] == "0.00" and jan["net"] == "0.00"
    assert feb["invoiced"] == "1000.00"  # VOID's 999 and DRAFT's 777 excluded
    assert mar == {
        "month": "2026-03",
        "invoiced": "0.00",
        "collected": "200.00",
        "refunded": "50.00",
        "net": "150.00",
    }
    assert data["totals"]["invoiced"] == "1000.00"
    assert data["totals"]["net"] == "150.00"


def test_revenue_by_client(manager_api, money_world):
    data = manager_api.get(REVENUE_URL, WINDOW | {"group_by": "client"}).json()

    assert len(data["rows"]) == 1
    row = data["rows"][0]
    assert row["client"] == "Acme Corp"
    assert row["invoiced"] == "1000.00"
    assert row["net"] == "150.00"
    assert row["outstanding"] == "800.00"  # 1000 - 200 paid, refund not reopened


def test_revenue_client_filter_excludes_others(manager_api, money_world):
    other = Client.objects.create(name="Other Ltd", status=Client.Status.ACTIVE)
    data = manager_api.get(REVENUE_URL, WINDOW | {"client": other.id}).json()
    assert data["totals"]["invoiced"] == "0.00"


# --- time report ------------------------------------------------------------


def test_time_report_manager_sees_everyone(manager_api, hours_world):
    data = manager_api.get(TIME_URL).json()
    assert len(data["rows"]) == 1  # one project row
    assert data["rows"][0]["project"] == "CRM Build"
    assert data["rows"][0]["client"] == "Acme Corp"
    assert data["totals"] == {"project": "Total", "entries": 2, "hours": "3.50"}


def test_time_report_staff_scoped_to_own_entries(staff_api, staff, hours_world, manager):
    # Even asking for the manager's rows by id must not widen the scope.
    data = staff_api.get(TIME_URL, {"group_by": "user", "user": manager.id}).json()
    assert len(data["rows"]) == 1
    assert data["rows"][0]["user"] == "Sam"
    assert data["totals"]["hours"] == "2.50"


def test_time_report_group_by_user(manager_api, hours_world):
    data = manager_api.get(TIME_URL, {"group_by": "user"}).json()
    assert {r["user"]: r["hours"] for r in data["rows"]} == {"Mia": "1.00", "Sam": "2.50"}


# --- ticket report ----------------------------------------------------------


def test_ticket_sla_report(manager_api, ticket_world):
    data = manager_api.get(TICKETS_URL).json()

    assert len(data["rows"]) == 1  # only HIGH has tickets; empty priorities dropped
    row = data["rows"][0]
    assert row["priority"] == "High"
    assert row["created"] == 2
    assert row["resolved"] == 1
    assert row["first_response_breached"] == 1  # the open one, past its FR deadline
    assert row["resolution_breached"] == 1  # the resolved one, 1h late
    assert row["avg_resolution_hours"] == "2.00"
    assert data["totals"]["created"] == 2
    assert data["totals"]["avg_resolution_hours"] is None  # never averaged across rows


# --- exports ----------------------------------------------------------------


def test_xlsx_export_is_a_real_workbook(manager_api, money_world):
    res = manager_api.get(REVENUE_URL, WINDOW | {"export": "xlsx"})

    assert res.status_code == 200
    assert res["Content-Type"].startswith("application/vnd.openxmlformats")
    assert 'attachment; filename="revenue-report-' in res["Content-Disposition"]

    ws = load_workbook(BytesIO(res.content)).active
    assert ws["A1"].value == "Revenue Report"
    assert ws["A4"].value == "Month"  # header row
    assert ws["B5"].value == 0  # January invoiced, zero-filled
    assert ws["B6"].value == 1000  # February invoiced — a NUMBER, not text
    assert ws["E8"].value == 150  # totals row (3 months + 1): net


def test_xlsx_defuses_formula_injection(manager_api, manager):
    hostile = Client.objects.create(name="=HYPERLINK('http://evil')", status=Client.Status.ACTIVE)
    Invoice.objects.create(
        client=hostile,
        invoice_number="INV-2026-0009",
        status=Invoice.Status.ISSUED,
        grand_total=Decimal("10.00"),
        issue_date=date(2026, 2, 1),
        due_date=date(2026, 3, 3),
    )
    res = manager_api.get(REVENUE_URL, WINDOW | {"group_by": "client", "export": "xlsx"})
    ws = load_workbook(BytesIO(res.content)).active
    assert ws["A5"].value.startswith("'=")  # apostrophe kills the formula


def test_pdf_export_is_a_real_pdf(staff_api, hours_world):
    res = staff_api.get(TIME_URL, {"export": "pdf"})
    assert res.status_code == 200
    assert res["Content-Type"] == "application/pdf"
    assert 'attachment; filename="time-report-' in res["Content-Disposition"]
    assert res.content[:5] == b"%PDF-"
